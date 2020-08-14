import datetime
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import re
from numpy import *
from openpyxl import Workbook
from openpyxl.comments import Comment
from bs4 import *
from urllib.request import urlopen as uReq
from bs4 import BeautifulSoup as soup
import numpy as np
import smtplib

today = datetime.datetime.today()

m_dic = {"OCAK": "JANUARY", "ŞUBAT": "FEBRUARY", "MART": "MARCH", "NİSAN": "APRIL",
         "MAYIS": "MAY", "HAZİRAN": "JUNE", "TEMMUZ": "JULY", "AĞUSTOS": "AUGUST",
         "EYLÜL": "SEPTEMBER", "EKİM": "OCTOBER", "KASIM": "NOVEMBER", "ARALIK": "DECEMBER", }

m_ndic = {1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan",
          5: "Mayıs", 6: "Haziran", 7: "Temmuz", 8: "Ağustos",
          9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"}


# Accessing Excel File
def load_file(path, sheet_name):
    # path = path of the your document
    # sheet_name =  which worksheet in the workbook

    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)

    ws = wb.get_sheet_by_name(sheet_name)

    return ws, wb


# Contract object consist of their NAME,NUMBER,TOTAL VAlUE,START DAY,END DAY,PRICE UPDATE PERIOD and UPDATE METHOD
class contract():

    def __init__(self, customer_name, contract_number, total_cont_value, start_d, end_d, periodic_pricing_period,
                 update_operation, method):
        self.customer_name = customer_name
        self.contract_number = contract_number
        self.total_cont_value = total_cont_value
        self.end_d = end_d
        self.start_d = start_d
        self.range_time = self.end_d - self.start_d
        self.elapsed_time = today - self.start_d
        self.HMDR = self.end_d - today
        self.periodic_pricing_period = periodic_pricing_period
        self.method = method
        self.update_operation = update_operation
        if self.periodic_pricing_period != None:
            self.periodic_pricing_status = True
            self.periodic_pricing_month = self.periodic_pricing_period.month
            self.periodic_pricing_year = self.periodic_pricing_period.year
        else:
            self.periodic_pricing_status = False
            self.periodic_pricing_month = None
            self.periodic_pricing_year = None
        if self.HMDR.days <= 90:
            self.Alert = True
        else:
            self.Alert = False

    def __str__(self):
        return (f" CUSTOMER NAME: {self.customer_name}  \n \
        CONTRACT NUMBER: {self.contract_number}  \n \
        CONTRACT START DATE:   {self.start_d}  \n \
        CONTRACT END DATE:  {self.end_d} \n \
        TOTAL CONTRACT VALUE: {self.total_cont_value} \n \
        HOW MANY DAYS REMAIN:{(self.HMDR.days)} \n \
        ALERT STATUS: {self.Alert}\n \
        CHANGING PERIOD: {self.periodic_pricing_year}\n \
        CHANGING MOUNT: {self.periodic_pricing_month}\n \
        UPGRADING OPERATION: {self.update_operation}\n \
        METHOD: {self.method}")


# Contracts object is joint all created contract by information from excel file
class Contracts():

    def __init__(self, ws):
        self.contract_list = []
        i = 4
        while ws.cell(row=i, column=1).value != None:
            self.new_conract = contract(ws.cell(row=i, column=3).value, ws.cell(row=i, column=4).value,
                                        ws.cell(row=i, column=6).value, ws.cell(row=i, column=12).value,
                                        ws.cell(row=i, column=13).value, ws.cell(row=i, column=26).value,
                                        ws.cell(row=i, column=27).value, ws.cell(row=i, column=28).value)
            # in this point you might have to arrange column index according to your excel sheet constuction
            self.contract_list.append(self.new_conract)
            i += 1

    def __len__(self):
        return len(self.contract_list)

    # alertedinspection function inspect is there any contract that has in alerted situation
    def alertedinspection(self):
        alertlisted = []
        for item in self.contract_list:
            if item.Alert == True:
                alertlisted.append(item)
            else:
                pass
        return alertlisted

    # changinginspection function inspect is there any contract that must be pricing updated
    def changinginspection(self):
        changinglist = []
        for item in self.contract_list:
            if (item.periodic_pricing_status == True) and (
                    (item.periodic_pricing_year * 12) + item.periodic_pricing_month < (
                    today.year * 12) + today.month) and item.update_operation != 'DONE':
                changinglist.append(item)
            else:
                pass
        return changinglist


# tuik_scrap_text function scrap data from selected web page
def tuik_scrap_text():
    list_URL = ["https://www.vergidegundem.com/tr/pb_yurt_ici_uretici_fiyat_endeksi",
                "https://www.vergidegundem.com/tr/pb_tuketici_fiyat_endeksi"]
    list_rate = []
    for URL in list_URL:
        result = uReq(URL)
        page_html = result.read()
        result.close()
        page_soup = soup(page_html, "html.parser")
        mainframe = page_soup.find('div', {"class": "pb-div"})
        sub_soup = mainframe.contents
        list2 = []
        arr_main = np.empty((0, 6), str)
        dict_array = {}
        for item in sub_soup:
            try:
                if item['class'] == ["pb-table-1"]:
                    subitems = item.find_all('tr')
                    for subitem in subitems:
                        subitemx2s = subitem.find_all('td')
                        for subitemx2 in subitemx2s:
                            list2.append((subitemx2.string).replace(',', '.'))
                        arr_main = np.append(arr_main, np.array([list2]), axis=0)
                        list2 = []
                    dict_array.update({arr_main.item(0, 0): arr_main})
                    arr_main = np.empty((0, 6), str)
            except:
                pass
        list_rate.append(dict_array)
    return list_rate


# ito_scrap_text function scrap data from selected web page
def ito_scrap_text():
    list_URL = [
        "https://bilgibankasi.ito.org.tr/tr/istatistik-verileri/istanbul-ucretler-gecinme/genel-indeksin-degisim-oranlari/yillik-ortalama-degisim?year=95",
        " https://bilgibankasi.ito.org.tr/tr/istatistik-verileri/istanbul-ucretler-gecinme/genel-indeksin-degisim-oranlari/yillik-ortalama-degisim?year=63-2"]
    list_rate = []
    for URL in list_URL:
        PATH = "C:\\Program Files (x86)\\chromedriver.exe"
        driver = webdriver.Chrome(PATH)
        driver.get(URL)
        driver.maximize_window()
        try:
            element = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CLASS_NAME, "table-3"))
            )
            text = element.text
            pattern = r'[\w]+ [\d]*[\.]*[\d]* [\d]*[\.]*[\d]* [\d]*[\.]*[\d]* [\d]*[\.]*[\d]* [\d]*[\.]*[\d]* [\d]*[\.]*[\d]* [\d]*[\.]*[\d]* [\d]*[\.]*[\d]* [\d]*[\.]*[\d]*[\s]?[\d]*[\.]*[\d]*'
            result = re.findall(pattern, text)
            arr_main = np.empty((0, 11), str)
            for item in result:
                list1 = item.split()
                if len(list1) == 10:
                    list1.append('')
                arr_main = np.append(arr_main, np.array([list1]), axis=0)
            list_rate.append(arr_main)
        except:
            pass

    return list_rate


# find the proper rate for pricing update operation
def scrap_rate(year, month, method):
    def select_rate_tuik(year, month):

        try:
            ufe_tuik = tuik_text[0][str(year)].item(month, 4)
            tufe_tuik = tuik_text[1][str(year)].item(month, 4)
        except:
            ufe_tuik = ""
            tufe_tuik = ""
        return ufe_tuik, tufe_tuik

    def select_rate_ito(year, month):
        b = list(map(lambda num: num, ito_text[0][0])).index(str(year))
        ufe_tuik = ito_text[0].item(month, b)
        tufe_tuik = ito_text[1].item(month, b)

        return ufe_tuik, tufe_tuik

    if method == "ITO":
        return select_rate_ito(year, month)
    elif method == "TUIK":
        return select_rate_tuik(year, month)


# match contract and thier rates
def find_rate(list_change):
    dict_rate = {}
    for item in list_change:
        item_rate = scrap_rate(item.periodic_pricing_year, item.periodic_pricing_month, item.method)
        dict_rate.update({item: item_rate})
    return dict_rate


# pricing update operation and the comment cell in the excel workbook

def update(rat_list, which_sheet):
    MR = which_sheet.max_row
    for item in rat_list.items():
        if item[1] != ('', ''):
            for Rrow in con_sheet.iter_rows(min_row=4, max_col=4, min_col=4, max_row=MR):
                for cell in Rrow:
                    if (cell.value == item[0].contract_number):
                        i = cell.row
                        print(item[0].total_cont_value)
                        inc_rate = (((float(item[1][0]) + float(item[1][1])) / 2) + 100) / 100
                        new_total_value = inc_rate * item[0].total_cont_value
                        print(new_total_value)
                        comment = Comment(f"New value must be {new_total_value}", "Script")
                        con_sheet.cell(row=i, column=6).comment = comment
                        print(con_sheet.cell(row=i, column=6).value)
                        print(con_sheet.cell(row=i, column=6).comment)

# create report and typing existing file
def report_text(selected_list, title='HEADER'):
    text_report = title + '\n' '------------------------------'
    for item in selected_list:
        text_report = text_report + '\n' + str(item.contract_number)
    return text_report

# Sending mail for notifying

def sendmail(To, S_TEXT, B_TEXT):
    smtp_object = smtplib.SMTP('smtp.gmail.com', 587)
    smtp_object.ehlo()
    smtp_object.starttls()
    email = 'your email'
    password = 'your password'
    smtp_object.login(email, password)
    from_address = email
    to_address = To
    subject = S_TEXT
    message = B_TEXT
    msg = "Subject: " + subject + '\n' + message

    smtp_object.sendmail(from_address, to_address, msg)


con_sheet, con_wb = load_file("your document path","your worksheet name")

tuik_text = tuik_scrap_text()

ito_text = ito_scrap_text()

a = Contracts(con_sheet)

print(len(a))

Alert_List = a.alertedinspection()

print(len(Alert_List))

Update_List = a.changinginspection()

print(len(Update_List))

Rate_Dic = find_rate(Update_List)

print(Rate_Dic)

update(Rate_Dic, con_sheet)

t1 = ('your existing I/O path for alerted list', 'ALERTED LIST', Alert_List)
t2 = ('your existing I/O path for update list', 'UPDATE LIST', Update_List)
list_item = [t1, t2]
for item in list_item:
    try:
        Item_Text = report_text(item[2], title=item[1])
        Item_Textbook = open(item[0], 'r', encoding="utf-8")
        current_item_list = Item_Textbook.read()
        print(current_item_list)
        print(Item_Text)
        if current_item_list != Item_Text:
            Item_Textbook = open(item[0], 'w', encoding="utf-8")
            Item_Textbook.write(Item_Text)
            sendmail('address that to sending', item[1], item[1] + ' has been changed')
        Item_Textbook.close()
    except:
        pass
con_wb.save("your document path")